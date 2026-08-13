#!/usr/bin/env python3
"""
Streamlit – generisanje PDF naloga.
Koristi fpdf2 + DejaVu (puna vidljivost, srpska slova, ista struktura podataka).
"""

import streamlit as st
import openpyxl
import pandas as pd
from fpdf import FPDF
from fpdf.enums import XPos, YPos
from datetime import datetime, timedelta
from pathlib import Path
import re
import zipfile
import io

APP_DIR = Path(__file__).parent
FONT_PATH = APP_DIR / "DejaVuSans.ttf"
FONT_BOLD = APP_DIR / "DejaVuSans-Bold.ttf"

st.set_page_config(page_title="Generisanje putnih naloga", page_icon="📄", layout="centered")
st.title("📄 Generisanje službenih putnih naloga")
st.markdown("Učitaj podatke iz **Google Sheets**-a ili Excel fajla i generiši PDF-ove.")


def parse_date(dstr):
    dstr = str(dstr).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(dstr, fmt)
        except ValueError:
            continue
    raise ValueError(f"Nepoznat format datuma: {dstr}")


def format_date(dt: datetime) -> str:
    return f"{dt.day:02d}.{dt.month:02d}.{dt.year}."


def format_amount(val) -> str:
    try:
        num = float(val)
        return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(val)


class NalogPDF(FPDF):
    def __init__(self):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.set_auto_page_break(auto=True, margin=12)
        self.add_font("DejaVu", "", str(FONT_PATH))
        self.add_font("DejaVu", "B", str(FONT_BOLD))

    def header_line(self, text, size=13):
        self.set_font("DejaVu", "B", size)
        self.cell(0, 7, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.ln(1)

    def label_value(self, label, value, label_w=52):
        self.set_font("DejaVu", "B", 9)
        self.cell(label_w, 5.5, label, new_x=XPos.RIGHT, new_y=YPos.TOP)
        self.set_font("DejaVu", "", 9)
        self.multi_cell(0, 5.5, str(value) if value else "")
        self.ln(0.5)

    def section_title(self, text):
        self.set_font("DejaVu", "B", 10)
        self.set_fill_color(235, 235, 235)
        self.cell(0, 6.5, text, border=1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        self.ln(1.5)


def create_nalog_bytes(worker: dict, day: datetime) -> bytes:
    pdf = NalogPDF()
    pdf.add_page()

    date_str = format_date(day)
    amount = format_amount(worker["dnevni_iznos"])
    prevoz = f"{worker['prevoz']}, {worker['registracija']}"
    org = worker["organizacija"]
    mesto = worker["mesto"]

    # ===== STRANA 1: NALOG =====
    pdf.header_line("NALOG ZA SLUŽBENO PUTOVANJE")
    pdf.set_font("DejaVu", "", 9)
    pdf.cell(0, 5, f"Organizacija: {org}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    pdf.section_title("PODACI O RADNIKU")
    pdf.label_value("Radnik:", worker["ime"])
    pdf.label_value("Radno mesto:", worker["pozicija"])
    pdf.label_value("Upućuje se na službeni put dana:", date_str)
    pdf.label_value("Sa zadatkom:", worker["zadatak"])
    pdf.label_value("Prevozno sredstvo:", prevoz)
    pdf.label_value("Dnevnica:", f"{amount} RSD")
    pdf.label_value("Zadržava se najdalje do:", date_str)
    pdf.label_value("Putni troškovi padaju na teret:", worker["teret"])
    pdf.label_value("Organizacija (teret):", org)
    pdf.ln(3)

    pdf.set_font("DejaVu", "", 8)
    pdf.multi_cell(0, 4,
        "U roku od 48 časova po povratku sa službenog puta i dolaska na posao, "
        "podneće pismeni izveštaj o obavljenom službenom poslu. "
        "Račun o učinjenim putnim troškovima podneti u roku od tri dana.")
    pdf.ln(4)

    pdf.set_font("DejaVu", "B", 9)
    pdf.cell(0, 5.5, "Odobravam isplatu akontacije u iznosu od dinara: _______________", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)
    pdf.cell(90, 5.5, "(M.P.)", new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.cell(0, 5.5, "Nalogodavac: ________________", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)

    # Putni račun
    pdf.section_title("PUTNI RAČUN")
    col_w = [28, 24, 24, 18, 18, 28, 40]
    headers = ["Datum", "Odlazak", "Povratak", "Km", "Dani", "Iznos", "Ukupno"]

    pdf.set_font("DejaVu", "B", 8)
    pdf.set_fill_color(240, 240, 240)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 5.5, h, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("DejaVu", "", 8)
    pdf.cell(col_w[0], 5.5, date_str, border=1, align="C")
    pdf.cell(col_w[1], 5.5, "07:00", border=1, align="C")
    pdf.cell(col_w[2], 5.5, "20:00", border=1, align="C")
    pdf.cell(col_w[3], 5.5, "13", border=1, align="C")
    pdf.cell(col_w[4], 5.5, "1", border=1, align="C")
    pdf.cell(col_w[5], 5.5, amount, border=1, align="R")
    pdf.cell(col_w[6], 5.5, amount, border=1, align="R")
    pdf.ln()

    for _ in range(2):
        for w in col_w:
            pdf.cell(w, 5.5, "", border=1)
        pdf.ln()

    pdf.ln(1)
    pdf.set_font("DejaVu", "B", 9)
    pdf.cell(0, 5.5, f"SVEGA: {amount} RSD", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="R")
    pdf.ln(1)
    pdf.set_font("DejaVu", "", 9)
    pdf.label_value("Primljena akontacija:", amount)
    pdf.label_value("Ostaje za isplatu/uplatu:", amount)
    pdf.label_value("Mesto i datum:", f"{mesto}, {date_str}")
    pdf.ln(4)

    pdf.set_font("DejaVu", "", 8)
    pdf.cell(60, 5.5, "Likvidirao: _______________", new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.cell(60, 5.5, "Rukovodilac: _______________", new_x=XPos.RIGHT, new_y=YPos.TOP)
    pdf.cell(0, 5.5, "Nalogodavac: _______________", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    # ===== STRANA 2: IZVEŠTAJ =====
    pdf.add_page()
    pdf.header_line("IZVEŠTAJ SA SLUŽBENOG PUTA")
    pdf.ln(2)

    pdf.set_font("DejaVu", "", 10)
    pdf.multi_cell(0, 5.5, "Službeni put protekao po planu i bez vanrednih događaja.")
    pdf.ln(3)

    pdf.set_font("DejaVu", "", 9)
    for _ in range(10):
        pdf.cell(0, 6.5, "_" * 95, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(6)
    pdf.set_font("DejaVu", "B", 9)
    pdf.cell(0, 5.5, f"Datum: {date_str}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 5.5, f"Radnik: {worker['ime']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(8)
    pdf.cell(0, 5.5, "Potpis radnika: ________________________", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


def load_workers_from_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Timovi" not in wb.sheetnames:
        st.error("U Excel fajlu ne postoji sheet **Timovi**.")
        return []
    ws = wb["Timovi"]
    workers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        workers.append({
            "ime": str(row[1]).strip(),
            "pozicija": str(row[2]).strip() if row[2] else "Službenik obezbeđenja",
            "datum_pocetka": row[3],
            "datum_kraja": row[4],
            "mesto": str(row[8]).strip() if len(row) > 8 and row[8] else "Zaječar",
            "zadatak": str(row[9]).strip() if len(row) > 9 and row[9] else "Obavljanje službenog zadatka na terenu",
            "prevoz": str(row[10]).strip() if len(row) > 10 and row[10] else "Ford Fokus 1.4",
            "registracija": str(row[11]).strip() if len(row) > 11 and row[11] else "ZA095LE",
            "dnevni_iznos": row[12] if len(row) > 12 and row[12] else 3471,
            "teret": str(row[14]).strip() if len(row) > 14 and row[14] else "poslodavca",
            "organizacija": str(row[15]).strip() if len(row) > 15 and row[15] else "Business Centre - Gold Guard DOO Zaječar",
        })
    return workers


def load_workers_from_google_sheets(sheet_url: str):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not match:
        st.error("Neispravan Google Sheets URL.")
        return []
    spreadsheet_id = match.group(1)
    csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid=0"
    try:
        df = pd.read_csv(csv_url)
    except Exception as e:
        st.error(f"Ne mogu da učitam Google Sheet. Proveri deljenje (Anyone with the link). Greška: {e}")
        return []
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {
        "ime": ["Član tima", "Ime", "Radnik", "Name"],
        "pozicija": ["Pozicija", "Radno mesto"],
        "datum_pocetka": ["Datum početka", "Datum pocetka", "Početak"],
        "datum_kraja": ["Datum kraja", "Kraj"],
        "mesto": ["Mesto", "Lokacija"],
        "zadatak": ["Zadatak", "Opis posla"],
        "prevoz": ["Prevozno sredstvo", "Prevoz"],
        "registracija": ["Registracija", "Reg. oznaka"],
        "dnevni_iznos": ["Dnevni iznos (RSD)", "Dnevni iznos", "Dnevnica"],
        "teret": ["Troškovi padaju na teret", "Teret"],
        "organizacija": ["Organizacija", "Firma"],
    }
    def find_col(names):
        for n in names:
            if n in df.columns: return n
        return None
    workers = []
    for _, row in df.iterrows():
        ime_col = find_col(col_map["ime"])
        if not ime_col or pd.isna(row.get(ime_col)): continue
        def get(key, default=""):
            col = find_col(col_map[key])
            if col and not pd.isna(row.get(col)): return str(row[col]).strip()
            return default
        workers.append({
            "ime": get("ime"),
            "pozicija": get("pozicija", "Službenik obezbeđenja"),
            "datum_pocetka": get("datum_pocetka"),
            "datum_kraja": get("datum_kraja"),
            "mesto": get("mesto", "Zaječar"),
            "zadatak": get("zadatak", "Obavljanje službenog zadatka na terenu"),
            "prevoz": get("prevoz", "Ford Fokus 1.4"),
            "registracija": get("registracija", "ZA095LE"),
            "dnevni_iznos": get("dnevni_iznos", 3471),
            "teret": get("teret", "poslodavca"),
            "organizacija": get("organizacija", "Business Centre - Gold Guard DOO Zaječar"),
        })
    return workers


def generate_pdfs(workers) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for w in workers:
            start = parse_date(w["datum_pocetka"])
            end = parse_date(w["datum_kraja"])
            current = start
            while current <= end:
                pdf_bytes = create_nalog_bytes(w, current)
                safe_name = re.sub(r"[^\w\s-]", "", w["ime"]).replace(" ", "_")
                filename = f"{safe_name}_{current.strftime('%Y%m%d')}.pdf"
                zipf.writestr(filename, pdf_bytes)
                current += timedelta(days=1)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ===== UI =====
tab1, tab2 = st.tabs(["📊 Google Sheets / Excel", "ℹ️ Uputstvo"])

with tab1:
    st.subheader("1. Izvor podataka")
    source = st.radio("Odakle učitavam podatke?", ["Upload Excel fajla (.xlsx)", "Google Sheets (javni link)"], horizontal=True)
    workers = []

    if source == "Upload Excel fajla (.xlsx)":
        uploaded = st.file_uploader("Izaberi Excel fajl", type=["xlsx"])
        if uploaded:
            workers = load_workers_from_excel(uploaded.getvalue())
            st.success(f"Učitano **{len(workers)}** radnika.")
    else:
        sheet_url = st.text_input("Nalepi link ka Google Sheets-u", placeholder="https://docs.google.com/spreadsheets/d/XXXX/edit")
        if sheet_url:
            with st.spinner("Učitavam..."):
                workers = load_workers_from_google_sheets(sheet_url)
            if workers:
                st.success(f"Učitano **{len(workers)}** radnika.")

    if workers:
        st.subheader("2. Pregled")
        st.dataframe(pd.DataFrame(workers)[["ime", "pozicija", "datum_pocetka", "datum_kraja", "mesto", "prevoz"]], use_container_width=True)
        total = sum((parse_date(w["datum_kraja"]) - parse_date(w["datum_pocetka"])).days + 1 for w in workers)
        st.info(f"Biće generisano **{total}** PDF naloga.")
        st.subheader("3. Generisanje")
        if st.button("🚀 GENERIŠI PDF NALOGE", type="primary", use_container_width=True):
            with st.spinner(f"Generišem {total} PDF-ova..."):
                try:
                    zip_bytes = generate_pdfs(workers)
                    st.success(f"✅ Generisano **{total}** PDF naloga!")
                    st.download_button("⬇️ Preuzmi ZIP", data=zip_bytes,
                        file_name=f"putni_nalozi_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                        mime="application/zip", use_container_width=True)
                except Exception as e:
                    st.error(str(e))
                    st.exception(e)

with tab2:
    st.markdown("""
**Struktura PDF-a (ista logika kao originalni obrazac):**
- Strana 1: Nalog za službeno putovanje + Putni račun
- Strana 2: Izveštaj sa službenog puta

Sva polja su potpuno vidljiva, sa podrškom za srpska slova.
""")

st.caption("Business Centre - Gold Guard")
